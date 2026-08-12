"""
把 v5 試算表的示範案例載入資料庫。

    python scripts/load_demo.py

**這是選配的，不是種子資料。** `import_seed.py` 匯的是官方參考資料（代碼表、
係數、熱值、GWP），任何使用者都需要；這一支匯的是「示範小吃店」——電號
`01-23-4567-89`、車牌 `3888-AB` 都是編的。假資料不該混進每個人的資料庫，
所以拆成兩支，要不要載入由你決定。

    python scripts/load_demo.py --clear     # 移除示範資料

載入的東西對應官方表單：

    表二_邊界設定    → Organization、BoundaryExclusion
    表三_排放源清冊  → EmissionSource（5 個排放源）
    活動數據登錄     → ActivityRecord（5 筆單據，只讀 A~J 欄）

**只讀 A~J 欄是刻意的。** K 欄之後是試算表用公式算的分攤與排放量，那正是
服務層要自己算的東西。把它們讀進來當輸入，就等於用試算表的答案驗試算表。
它們只在測試裡當對照組（見 tests/test_service.py）。

讀取邏輯放在這裡而不是 app/，跟 extract_codes.py 同一個理由：app/ 是會被
應用程式載入的程式碼，示範資料的解析不屬於那裡。
"""

from __future__ import annotations

import argparse
import datetime as dt
import sys
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl                                                  # noqa: E402
from sqlalchemy import select                                    # noqa: E402
from sqlalchemy.orm import Session                               # noqa: E402

from app.db import SessionLocal, init_db                         # noqa: E402
from app.models import (                                         # noqa: E402
    ActivityRecord, BoundaryExclusion, DataQuality, DirectIndirect,
    EmissionResult, EmissionSource, EmissionType, EvidenceType, Organization,
)
from app.seed import DEFAULT_WORKBOOK                            # noqa: E402

SHEET_BOUNDARY = "表二_邊界設定"
SHEET_SOURCES = "表三_排放源清冊"
SHEET_ACTIVITY = "活動數據登錄"

# 試算表用這些字樣表示「這一格還沒填」，不是真的內容。
_PLACEHOLDERS = {"（待填）", "(待填)", "（自行補充）", "(自行補充)", "—", "-", "－"}


class DemoFormatError(ValueError):
    """示範分頁的版面與預期不符。"""


@dataclass(frozen=True)
class DemoSource:
    source_no: str
    name: str
    process_no: str | None
    process_code: str | None
    equipment_no: str | None
    equipment_code: str | None
    material_code: str | None
    is_biomass: bool
    direct_indirect: str
    emission_type: str
    produces_co2: bool
    produces_ch4: bool
    produces_n2o: bool
    factor_key: str
    note: str | None


@dataclass(frozen=True)
class DemoRecord:
    seq: int
    source_name: str
    evidence_type: str
    evidence_ref: str
    period_start: dt.date
    period_end: dt.date
    raw_quantity: float
    unit: str
    data_quality: str
    estimation_basis: str | None
    # 試算表 V 欄自己算的 tCO2e。不當輸入，只在測試裡當對照組。
    spreadsheet_tco2e: float | None


@dataclass(frozen=True)
class DemoCase:
    org_name: str
    reporting_year_roc: int
    base_year_roc: int | None
    boundary_method: str | None
    county: str | None
    district: str | None
    postal_code: str | None
    exclusions: list[tuple[str, str]]
    sources: list[DemoSource]
    records: list[DemoRecord]


# --------------------------------------------------------------------------
# 讀取
# --------------------------------------------------------------------------

def _text(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return None if not s or s in _PLACEHOLDERS else s


def _required(value, where: str) -> str:
    s = _text(value)
    if s is None:
        raise DemoFormatError(f"{where} 不可為空")
    return s


def _date(value, where: str) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise DemoFormatError(f"{where} 不是日期：{value!r}")


def _flag(value) -> bool:
    """表三的 v / 空白，以及 是 / 否。"""
    s = _text(value)
    return s is not None and s.lower() in {"v", "是", "y", "yes", "true"}


def _check(ws, row: int, expected: dict[str, str]) -> None:
    for column, want in expected.items():
        got = _text(ws[f"{column}{row}"].value)
        first = got.splitlines()[0].strip() if got else None
        if first != want:
            raise DemoFormatError(
                f"分頁「{ws.title}」{column}{row} 應為「{want}」，實際為「{got}」。"
                f"版面可能被更動過。"
            )


def _read_boundary(wb) -> dict:
    ws = wb[SHEET_BOUNDARY]
    _check(ws, 5, {"A": "事業名稱"})
    _check(ws, 8, {"A": "盤查年度"})

    year = _text(ws["B8"].value)
    if year is None or not year.isdigit():
        raise DemoFormatError(f"{SHEET_BOUNDARY}!B8 盤查年度應為民國年數字，實際為「{year}」")

    base = _text(ws["B9"].value)
    exclusions = []
    for r in range(21, 30):
        item, reason = _text(ws[f"B{r}"].value), _text(ws[f"C{r}"].value)
        if item is None:
            continue
        if reason is None:
            # 表二 C 欄標題就寫著「排除理由（必須寫，不可留白）」。
            # 沒有理由的排除項目正是查驗會挑的東西，不能讓它靜靜通過。
            raise DemoFormatError(
                f"{SHEET_BOUNDARY} 第 {r} 列「{item}」沒有填排除理由"
            )
        exclusions.append((item, reason))

    return {
        "org_name": _required(ws["B5"].value, f"{SHEET_BOUNDARY}!B5 事業名稱"),
        "reporting_year_roc": int(year),
        "base_year_roc": int(base) if base and base.isdigit() else None,
        "boundary_method": _text(ws["B11"].value),
        "county": _text(ws["B12"].value),
        "district": _text(ws["B13"].value),
        "postal_code": _text(ws["B14"].value),
        "exclusions": exclusions,
    }


def _read_sources(wb) -> list[DemoSource]:
    ws = wb[SHEET_SOURCES]
    _check(ws, 3, {
        "A": "排放源編號", "C": "製程編號", "D": "製程代碼", "F": "設備編號",
        "G": "設備代碼", "I": "原燃物料代碼", "L": "直接/間接", "M": "排放型式",
        "Q": "對應係數編號",
    })

    # 資料區是連續的：第 4 列開始，遇到整列空白就結束。之後還有一列說明文字
    # （「完整性檢查：清冊上每一個排放源…」），它 A 欄有字但 B、Q 皆空。
    #
    # 為什麼不用「A 欄有值就當資料」：那會把說明文字讀成排放源。
    # 為什麼不用「B 欄有值才算」：那會讓漏填名稱的排放源靜靜消失，而漏掉
    # 一個排放源正是這張表最該防的事。改用空白列當終點，再回頭檢查後面還有
    # 沒有漏網的 —— 有的話代表清單中間被插了空白列，資料會被截斷。
    last_row = ws.max_row
    for r in range(4, ws.max_row + 1):
        if all(_text(ws.cell(r, c).value) is None for c in range(1, 19)):
            last_row = r - 1
            break

    for r in range(last_row + 1, ws.max_row + 1):
        stray = _text(ws[f"A{r}"].value)
        if stray and _text(ws[f"Q{r}"].value):
            raise DemoFormatError(
                f"{SHEET_SOURCES} 第 {r} 列還有排放源「{stray}」，但它前面有空白列，"
                f"清單會被截斷。請把空白列刪掉，讓排放源連續排列。"
            )

    out: list[DemoSource] = []
    for r in range(4, last_row + 1):
        source_no = _text(ws[f"A{r}"].value)
        if source_no is None:
            continue
        where = f"{SHEET_SOURCES} 第 {r} 列（{source_no}）"
        out.append(DemoSource(
            source_no=source_no,
            name=_required(ws[f"B{r}"].value, f"{where} 排放源名稱"),
            process_no=_text(ws[f"C{r}"].value),
            process_code=_text(ws[f"D{r}"].value),
            equipment_no=_text(ws[f"F{r}"].value),
            equipment_code=_text(ws[f"G{r}"].value),
            material_code=_text(ws[f"I{r}"].value),
            is_biomass=_flag(ws[f"K{r}"].value),
            direct_indirect=_required(ws[f"L{r}"].value, f"{where} 直接/間接"),
            emission_type=_required(ws[f"M{r}"].value, f"{where} 排放型式"),
            produces_co2=_flag(ws[f"N{r}"].value),
            produces_ch4=_flag(ws[f"O{r}"].value),
            produces_n2o=_flag(ws[f"P{r}"].value),
            factor_key=_required(ws[f"Q{r}"].value, f"{where} 對應係數編號"),
            note=_text(ws[f"R{r}"].value),
        ))

    if not out:
        raise DemoFormatError(f"分頁「{SHEET_SOURCES}」讀不到任何排放源")
    return out


def _read_records(wb) -> list[DemoRecord]:
    """
    只讀 A~J 欄（單據內容），K 欄之後是試算表的公式，那是服務層要自己算的。

    這張表有三種列，只有第一種是資料：

        A 有序號、B 與 G 有值    真正的單據          第 7~11 列
        A 有序號、B 與 G 空白    預先編號的空白範本   第 12~30 列
        A 空白、B 有名稱         分頁自己的小計列     第 33~35 列（範疇一／二／合計）

    只看「A 是不是數字」會把空白範本讀進來（得到一堆空記錄）；只看「B 有沒有
    值」會把小計列讀進來（總量直接翻倍，而且不會有錯誤訊息）。兩個條件都要。
    """
    ws = wb[SHEET_ACTIVITY]
    _check(ws, 6, {
        "A": "序號", "B": "排放源名稱", "C": "佐證類型", "D": "佐證編號",
        "E": "期間起", "F": "期間迄", "G": "原始活動數據", "H": "單位",
        "I": "資料品質",
    })

    out: list[DemoRecord] = []
    for r in range(7, ws.max_row + 1):
        seq = _text(ws[f"A{r}"].value)
        name = _text(ws[f"B{r}"].value)
        qty = ws[f"G{r}"].value

        if seq is None or not seq.isdigit():
            continue                      # 小計列
        if name is None or qty is None:
            continue                      # 預先編號的空白範本列

        where = f"{SHEET_ACTIVITY} 第 {r} 列（序號 {seq}）"
        quality = _required(ws[f"I{r}"].value, f"{where} 資料品質")
        basis = _text(ws[f"J{r}"].value)

        v = ws[f"V{r}"].value
        out.append(DemoRecord(
            seq=int(seq),
            source_name=name,
            evidence_type=_required(ws[f"C{r}"].value, f"{where} 佐證類型"),
            evidence_ref=_required(ws[f"D{r}"].value, f"{where} 佐證編號"),
            period_start=_date(ws[f"E{r}"].value, f"{where} 期間起"),
            period_end=_date(ws[f"F{r}"].value, f"{where} 期間迄"),
            raw_quantity=float(qty),
            unit=_required(ws[f"H{r}"].value, f"{where} 單位"),
            data_quality=quality,
            estimation_basis=basis,
            spreadsheet_tco2e=float(v) if isinstance(v, (int, float)) else None,
        ))

    if not out:
        raise DemoFormatError(f"分頁「{SHEET_ACTIVITY}」讀不到任何單據")
    return out


def read_demo(path: str | Path = DEFAULT_WORKBOOK) -> DemoCase:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"找不到試算表：{path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        missing = {SHEET_BOUNDARY, SHEET_SOURCES, SHEET_ACTIVITY} - set(wb.sheetnames)
        if missing:
            raise DemoFormatError(f"試算表缺少分頁：{'、'.join(sorted(missing))}")

        boundary = _read_boundary(wb)
        sources = _read_sources(wb)
        records = _read_records(wb)

        known = {s.name for s in sources}
        unknown = {r.source_name for r in records} - known
        if unknown:
            raise DemoFormatError(
                f"活動數據登錄用到表三沒有的排放源：{'、'.join(sorted(unknown))}。"
                f"名稱是兩張表唯一的連結，錯一個字就對不起來。"
            )

        return DemoCase(sources=sources, records=records, **boundary)
    finally:
        wb.close()


# --------------------------------------------------------------------------
# 寫入
# --------------------------------------------------------------------------

def load_demo(session: Session, demo: DemoCase) -> Organization:
    """
    寫進資料庫，依自然鍵 upsert，可重複執行。

    自然鍵：Organization 用 (name, reporting_year_roc)、EmissionSource 用
    (org_id, source_no)、ActivityRecord 用 (source_id, evidence_ref)。
    """
    org = session.scalars(
        select(Organization).where(
            Organization.name == demo.org_name,
            Organization.reporting_year_roc == demo.reporting_year_roc,
        )
    ).one_or_none()

    if org is None:
        org = Organization(
            name=demo.org_name, reporting_year_roc=demo.reporting_year_roc)
        session.add(org)

    org.base_year_roc = demo.base_year_roc
    org.boundary_method = demo.boundary_method
    org.county = demo.county
    org.district = demo.district
    org.postal_code = demo.postal_code
    session.flush()

    existing_ex = {
        e.excluded_item: e for e in session.scalars(
            select(BoundaryExclusion).where(BoundaryExclusion.org_id == org.id)).all()
    }
    for item, reason in demo.exclusions:
        found = existing_ex.get(item)
        if found is None:
            session.add(BoundaryExclusion(
                org_id=org.id, excluded_item=item, reason=reason))
        else:
            found.reason = reason

    existing_src = {
        s.source_no: s for s in session.scalars(
            select(EmissionSource).where(EmissionSource.org_id == org.id)).all()
    }
    by_name: dict[str, EmissionSource] = {}
    for row in demo.sources:
        src = existing_src.get(row.source_no)
        if src is None:
            src = EmissionSource(org_id=org.id, source_no=row.source_no, name=row.name)
            session.add(src)
        src.name = row.name
        src.process_no = row.process_no
        src.process_code = row.process_code
        src.equipment_no = row.equipment_no
        src.equipment_code = row.equipment_code
        src.material_code = row.material_code
        src.is_biomass = row.is_biomass
        src.direct_indirect = DirectIndirect(row.direct_indirect)
        src.emission_type = EmissionType(row.emission_type)
        src.produces_co2 = row.produces_co2
        src.produces_ch4 = row.produces_ch4
        src.produces_n2o = row.produces_n2o
        src.factor_key = row.factor_key
        src.note = row.note
        by_name[row.name] = src
    session.flush()

    existing_rec = {
        (r.source_id, r.evidence_ref): r for r in session.scalars(
            select(ActivityRecord).join(EmissionSource).where(
                EmissionSource.org_id == org.id)).all()
    }
    for row in demo.records:
        src = by_name[row.source_name]
        rec = existing_rec.get((src.id, row.evidence_ref))
        if rec is None:
            rec = ActivityRecord(
                source_id=src.id, evidence_ref=row.evidence_ref,
                period_start=row.period_start, period_end=row.period_end,
                raw_quantity=row.raw_quantity, unit=row.unit,
            )
            session.add(rec)
        rec.period_start = row.period_start
        rec.period_end = row.period_end
        rec.raw_quantity = row.raw_quantity
        rec.unit = row.unit
        rec.data_quality = DataQuality(row.data_quality)
        rec.estimation_basis = row.estimation_basis
        rec.evidence_type = EvidenceType(row.evidence_type)
        rec.data_source = row.evidence_type

    session.flush()
    return org


def clear_demo(session: Session, demo: DemoCase) -> int:
    """移除示範資料。連同算出來的結果一起刪，否則會留下孤兒。"""
    org = session.scalars(
        select(Organization).where(
            Organization.name == demo.org_name,
            Organization.reporting_year_roc == demo.reporting_year_roc,
        )
    ).one_or_none()
    if org is None:
        return 0

    sources = session.scalars(
        select(EmissionSource).where(EmissionSource.org_id == org.id)).all()
    removed = 0
    for src in sources:
        for rec in session.scalars(
                select(ActivityRecord).where(ActivityRecord.source_id == src.id)).all():
            for res in session.scalars(
                    select(EmissionResult).where(EmissionResult.record_id == rec.id)).all():
                session.delete(res)
            session.delete(rec)
            removed += 1
        session.delete(src)
    for ex in session.scalars(
            select(BoundaryExclusion).where(BoundaryExclusion.org_id == org.id)).all():
        session.delete(ex)
    session.delete(org)
    return removed


# --------------------------------------------------------------------------

def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="載入 v5 試算表的示範案例")
    parser.add_argument("workbook", nargs="?", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--clear", action="store_true", help="移除示範資料")
    args = parser.parse_args(argv[1:])

    try:
        demo = read_demo(args.workbook)
    except (FileNotFoundError, DemoFormatError) as exc:
        print(f"讀取失敗：{exc}", file=sys.stderr)
        return 1

    init_db()
    session = SessionLocal()
    try:
        if args.clear:
            n = clear_demo(session, demo)
            session.commit()
            print(f"已移除示範案例「{demo.org_name}」（{n} 筆活動數據）")
            return 0

        org = load_demo(session, demo)
        session.commit()
        print(f"示範案例：{org.name}　{org.reporting_year_roc} 年度"
              f"（{org.year_start} ~ {org.year_end}）")
        print(f"  邊界排除   {len(demo.exclusions)} 項")
        print(f"  排放源     {len(demo.sources)} 個")
        print(f"  活動數據   {len(demo.records)} 筆")
        print()
        for row in demo.records:
            print(f"    {row.seq}. {row.source_name:<22} "
                  f"{row.raw_quantity:>8,.1f} {row.unit:<5} {row.data_quality}")
        print()
        print("尚未計算排放量。跑 python scripts/calc_demo.py 或用 app/service.py。")
        return 0
    except Exception as exc:                     # noqa: BLE001
        session.rollback()
        print(f"載入失敗，已回滾：{exc}", file=sys.stderr)
        return 1
    finally:
        session.close()


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
