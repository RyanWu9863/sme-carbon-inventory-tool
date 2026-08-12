"""
種子資料讀取 — 從 碳盤查試算表_v5.xlsx 讀出係數、熱值、電力係數、GWP 與代碼表。

為什麼從試算表讀，而不是手抄成 Python 常數：

    手抄會產生第二份真相。係數改版時要記得改兩個地方，漏掉一個，
    程式與試算表就開始各說各話，而且不會有任何錯誤訊息 —— 兩邊都
    「正常執行」，只是答案不同。試算表已經是測試的基準（見
    tests/test_calculator.py），讓它同時當種子資料的來源，全專案
    就只有一份數字。

本模組只負責「讀」，不寫資料庫、不寫檔案 —— 因此可以單獨測試，
寫入的部分在 scripts/import_seed.py。這跟 calculator.py 是同一個
原則：把會失敗的邏輯留在純函式裡。

試算表若被改壞（欄位搬動、公式沒重算），這裡一律明確拋錯，不猜、
不填預設值。種子資料錯了整份報告都是錯的，而且錯得很安靜。
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# 試算表放在專案根目錄。用 __file__ 推導，所以專案資料夾搬到哪都能跑。
DEFAULT_WORKBOOK = Path(__file__).resolve().parents[1] / "碳盤查試算表_v5.xlsx"

# 試算表分頁名稱。改名的話這裡要跟著改，而不是讓程式去猜。
SHEET_FUEL = "燃料係數計算"
SHEET_ELECTRICITY = "電力係數"
SHEET_GWP = "GWP表"
SHEET_CODE = "代碼表"
SHEET_SOURCES = "資料來源與限制"

# 「燃料係數計算」分頁的版面：標題在第 6 列，資料從第 7 列開始。
FUEL_HEADER_ROW = 6
FUEL_FIRST_ROW = 7

_USAGE_MAP = {"固定燃燒": "stationary", "移動燃燒": "mobile"}

# 車輛技術別欄位在固定燃燒列填「—」表示不適用
_NOT_APPLICABLE = {"—", "-", "－", ""}


class SeedFormatError(ValueError):
    """試算表版面與預期不符。與其讀出半套資料，不如在這裡停下來。"""


@dataclass(frozen=True)
class FuelRow:
    """「燃料係數計算」的一列：公告係數與熱值假設併在同一列。"""

    factor_code: str          # TW-F-NG-S
    activity_key: str         # natural_gas_stationary
    display_name: str         # 天然氣
    usage: str                # stationary / mobile
    vehicle_tech: str | None  # 氧化觸媒；固定燃燒為 None
    unit: str                 # 公升 / 立方公尺
    co2_kg_per_tj: float
    ch4_kg_per_tj: float
    n2o_kg_per_tj: float
    ch4_gwp_gas: str          # 甲烷 or 石化甲烷
    source_ref: str           # 附表一-固定 第45列
    material_code: str        # 050002
    heating_value_kcal: float
    heating_value_source: str
    heating_value_version: str | None
    keywords: str | None
    note: str | None
    # 試算表 U 欄自己算出來的合計係數。不用於計算，只用於比對：
    # 程式重算的結果必須與它一致，否則兩邊有一邊錯了。
    spreadsheet_total: float


@dataclass(frozen=True)
class ElectricityRow:
    factor_code: str
    display_name: str
    kgco2e_per_kwh: float
    year_roc: int
    unit: str
    source: str
    note: str | None


@dataclass(frozen=True)
class GwpRow:
    standard: str
    gas_name: str
    formula: str | None
    gwp100: float
    note: str | None


@dataclass(frozen=True)
class CodeRow:
    code: str
    name: str


@dataclass(frozen=True)
class FactorSource:
    """
    這批係數出自哪一份公告。

    FactorSet 靠它回答「這個數字是哪一版算的」—— models.py 的第三個設計
    決策（計算結果做快照）沒有這些欄位就落空了。

    一樣不手抄：出處寫在試算表「資料來源與限制」B4，那是報告要引用的
    同一段文字，抄一份到 Python 常數裡遲早會跟它對不上。
    """

    announcement: str          # 溫室氣體排放係數
    doc_no: str                # 環部授氣字第1139101231號
    publish_date: dt.date      # 2024-02-05
    publish_date_roc: str      # 113年2月5日
    description: str           # B4 原文，整段留著供報告引用


@dataclass(frozen=True)
class SeedData:
    kcal_to_tj: float
    gwp_standard: str
    factor_source: FactorSource
    fuels: list[FuelRow]
    electricity: list[ElectricityRow]
    gwp: list[GwpRow]
    process_codes: list[CodeRow]
    equipment_codes: list[CodeRow]
    material_codes: list[CodeRow]
    # 電力係數分頁列了年度但還沒填值的（例如 114 年待公告）。
    # 不是錯誤，但呼叫端該知道，否則盤查該年度時會查無係數。
    electricity_pending_years: list[int]


def _text(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _required_text(value, where: str) -> str:
    s = _text(value)
    if s is None:
        raise SeedFormatError(f"{where} 不可為空")
    return s


def _required_float(value, where: str) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise SeedFormatError(f"{where} 不可為空")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise SeedFormatError(f"{where} 不是數字：{value!r}") from exc


def _check_header(ws, row: int, expected: dict[str, str]) -> None:
    """欄位搬動過就停下來。讀錯欄位的種子資料看起來完全正常。"""
    for column, want in expected.items():
        got = _text(ws[f"{column}{row}"].value)
        # 標題含換行（CO2\nkg/TJ），只比對第一行
        got_first = got.splitlines()[0].strip() if got else None
        if got_first != want:
            raise SeedFormatError(
                f"分頁「{ws.title}」{column}{row} 應為「{want}」，實際為「{got}」。"
                f"欄位順序可能被更動過，請確認試算表版本。"
            )


def _read_fuels(wb) -> tuple[list[FuelRow], float, str]:
    ws = wb[SHEET_FUEL]

    _check_header(ws, FUEL_HEADER_ROW, {
        "A": "係數編號", "B": "activity_key", "C": "燃料名稱", "D": "用途",
        "E": "車輛技術別", "F": "活動單位", "G": "CO2", "H": "CH4", "I": "N2O",
        "J": "甲烷GWP類別", "K": "公告出處", "L": "原燃物料代碼", "M": "熱值",
        "U": "合計",
    })

    kcal_to_tj = _required_float(ws["C2"].value, f"{SHEET_FUEL}!C2 換算常數")
    gwp_standard = _required_text(ws["C3"].value, f"{SHEET_FUEL}!C3 GWP 版本")

    rows: list[FuelRow] = []
    for r in range(FUEL_FIRST_ROW, ws.max_row + 1):
        factor_code = _text(ws[f"A{r}"].value)
        if factor_code is None:
            continue          # 表格下方的說明列

        where = f"{SHEET_FUEL} 第 {r} 列（{factor_code}）"

        usage_label = _required_text(ws[f"D{r}"].value, f"{where} 用途")
        if usage_label not in _USAGE_MAP:
            raise SeedFormatError(
                f"{where} 用途為「{usage_label}」，只接受 {'／'.join(_USAGE_MAP)}"
            )

        tech = _text(ws[f"E{r}"].value)
        if tech in _NOT_APPLICABLE:
            tech = None

        total = ws[f"U{r}"].value
        if total is None:
            raise SeedFormatError(
                f"{where} U 欄（合計係數）沒有值。試算表的公式尚未計算 —— "
                f"用 Excel 或 LibreOffice 開啟後存檔，讓公式算出結果再匯入。"
            )

        rows.append(FuelRow(
            factor_code=factor_code,
            activity_key=_required_text(ws[f"B{r}"].value, f"{where} activity_key"),
            display_name=_required_text(ws[f"C{r}"].value, f"{where} 燃料名稱"),
            usage=_USAGE_MAP[usage_label],
            vehicle_tech=tech,
            unit=_required_text(ws[f"F{r}"].value, f"{where} 活動單位"),
            co2_kg_per_tj=_required_float(ws[f"G{r}"].value, f"{where} CO2"),
            ch4_kg_per_tj=_required_float(ws[f"H{r}"].value, f"{where} CH4"),
            n2o_kg_per_tj=_required_float(ws[f"I{r}"].value, f"{where} N2O"),
            ch4_gwp_gas=_required_text(ws[f"J{r}"].value, f"{where} 甲烷GWP類別"),
            source_ref=_required_text(ws[f"K{r}"].value, f"{where} 公告出處"),
            material_code=_required_text(ws[f"L{r}"].value, f"{where} 原燃物料代碼"),
            heating_value_kcal=_required_float(ws[f"M{r}"].value, f"{where} 熱值"),
            heating_value_source=_required_text(ws[f"N{r}"].value, f"{where} 熱值來源"),
            heating_value_version=_text(ws[f"O{r}"].value),
            keywords=_text(ws[f"V{r}"].value),
            note=_text(ws[f"W{r}"].value),
            spreadsheet_total=float(total),
        ))

    if not rows:
        raise SeedFormatError(f"分頁「{SHEET_FUEL}」讀不到任何燃料列")
    return rows, kcal_to_tj, gwp_standard


def _read_electricity(wb) -> tuple[list[ElectricityRow], list[int]]:
    ws = wb[SHEET_ELECTRICITY]
    _check_header(ws, 3, {
        "A": "係數編號", "B": "顯示名稱", "C": "係數值",
        "D": "活動單位", "H": "適用年度",
    })

    rows: list[ElectricityRow] = []
    pending: list[int] = []
    for r in range(4, ws.max_row + 1):
        code = _text(ws[f"A{r}"].value)
        if code is None:
            continue
        where = f"{SHEET_ELECTRICITY} 第 {r} 列（{code}）"
        year = int(_required_float(ws[f"H{r}"].value, f"{where} 適用年度"))

        # 尚未公告的年度（如 114）留空是預期狀態，不是錯誤。
        # 但要記下來 —— 盤查該年度時會查無係數，呼叫端得知道原因。
        if _text(ws[f"C{r}"].value) is None:
            pending.append(year)
            continue

        rows.append(ElectricityRow(
            factor_code=code,
            display_name=_required_text(ws[f"B{r}"].value, f"{where} 顯示名稱"),
            kgco2e_per_kwh=_required_float(ws[f"C{r}"].value, f"{where} 係數值"),
            year_roc=year,
            unit=_required_text(ws[f"D{r}"].value, f"{where} 活動單位"),
            source=_required_text(ws[f"G{r}"].value, f"{where} 資料來源"),
            note=_text(ws[f"I{r}"].value),
        ))

    if not rows:
        raise SeedFormatError(f"分頁「{SHEET_ELECTRICITY}」讀不到任何有值的係數")
    return rows, pending


def _read_gwp(wb) -> list[GwpRow]:
    ws = wb[SHEET_GWP]
    _check_header(ws, 3, {
        "A": "版本", "B": "氣體名稱", "C": "化學式", "D": "GWP100",
    })

    rows: list[GwpRow] = []
    for r in range(4, ws.max_row + 1):
        standard = _text(ws[f"A{r}"].value)
        gas = _text(ws[f"B{r}"].value)
        if standard is None or gas is None:
            continue          # 表格下方的說明列
        where = f"{SHEET_GWP} 第 {r} 列（{gas}）"
        rows.append(GwpRow(
            standard=standard,
            gas_name=gas,
            formula=_text(ws[f"C{r}"].value),
            gwp100=_required_float(ws[f"D{r}"].value, f"{where} GWP100"),
            note=_text(ws[f"E{r}"].value),
        ))

    if not rows:
        raise SeedFormatError(f"分頁「{SHEET_GWP}」讀不到任何 GWP 值")
    return rows


def _read_code_block(ws, code_col: str, name_col: str) -> list[CodeRow]:
    rows: list[CodeRow] = []
    for r in range(5, ws.max_row + 1):
        code = _text(ws[f"{code_col}{r}"].value)
        name = _text(ws[f"{name_col}{r}"].value)
        if code is None or name is None:
            continue
        rows.append(CodeRow(code=code, name=name))
    return rows


def _read_codes(wb) -> tuple[list[CodeRow], list[CodeRow], list[CodeRow]]:
    ws = wb[SHEET_CODE]
    _check_header(ws, 4, {
        "A": "製程代碼", "D": "設備代碼", "G": "原燃物料代碼",
    })
    # 三個代碼區塊並排在同一張表，各佔兩欄，長度不一。
    return (
        _read_code_block(ws, "A", "B"),
        _read_code_block(ws, "D", "E"),
        _read_code_block(ws, "G", "H"),
    )


def _read_factor_source(wb) -> FactorSource:
    """
    從「資料來源與限制」B4 讀出公告名稱、文號與公告日期。

    B4 是一段給人看的敘述：

        環境部 113年2月5日 環部授氣字第1139101231號公告「溫室氣體排放係數」
        附表一。法源：溫室氣體排放量盤查登錄及查驗管理辦法第4條第2項第1款。

    這裡從裡面挑出三樣結構化資料，整段原文也一併留著（description）——
    報告要引用的是原文，資料庫要查詢的是結構化欄位，兩個都要。

    抓不到就報錯。抓不到而給預設值，等於讓資料庫宣稱一份不存在的公告。
    """
    ws = wb[SHEET_SOURCES]
    _check_header(ws, 4, {"A": "燃料排放係數"})

    text = _required_text(ws["B4"].value, f"{SHEET_SOURCES}!B4 燃料排放係數來源")
    flat = " ".join(text.split())

    announcement = re.search(r"公告「([^」]+)」", flat)
    doc_no = re.search(r"(環部授氣字第\d+號)", flat)
    date = re.search(r"(\d{2,3})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", flat)

    missing = [
        label for label, m in
        (("公告名稱（「」框起來的部分）", announcement),
         ("公告文號（環部授氣字第…號）", doc_no),
         ("公告日期（民國年月日）", date))
        if m is None
    ]
    if missing:
        raise SeedFormatError(
            f"{SHEET_SOURCES}!B4 找不到 {'、'.join(missing)}。\n"
            f"實際內容：{flat[:120]}…\n"
            f"這段文字是 FactorSet 的版本依據，格式改了就不能照原樣解讀。"
        )

    roc_year, month, day = (int(g) for g in date.groups())
    try:
        publish_date = dt.date(roc_year + 1911, month, day)
    except ValueError as exc:
        raise SeedFormatError(
            f"{SHEET_SOURCES}!B4 的公告日期 {roc_year}年{month}月{day}日 不是有效日期"
        ) from exc

    return FactorSource(
        announcement=announcement.group(1),
        doc_no=doc_no.group(1),
        publish_date=publish_date,
        publish_date_roc=f"{roc_year}年{month}月{day}日",
        description=flat,
    )


def read_seed(path: str | Path) -> SeedData:
    """讀取整份試算表。任何版面異常都在這裡拋 SeedFormatError。"""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"找不到試算表：{path}\n"
            f"種子資料以 碳盤查試算表_v5.xlsx 為唯一來源，請確認檔案在專案根目錄。"
        )

    # data_only=True 取公式的計算結果（U 欄的合計係數）。
    # 若試算表從未被 Excel 開啟過，快取值會是 None，_read_fuels 會明確報錯。
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        required = {SHEET_FUEL, SHEET_ELECTRICITY, SHEET_GWP, SHEET_CODE, SHEET_SOURCES}
        missing = required - set(wb.sheetnames)
        if missing:
            raise SeedFormatError(f"試算表缺少分頁：{'、'.join(sorted(missing))}")

        fuels, kcal_to_tj, gwp_standard = _read_fuels(wb)
        electricity, pending = _read_electricity(wb)
        process_codes, equipment_codes, material_codes = _read_codes(wb)

        return SeedData(
            kcal_to_tj=kcal_to_tj,
            gwp_standard=gwp_standard,
            factor_source=_read_factor_source(wb),
            fuels=fuels,
            electricity=electricity,
            gwp=_read_gwp(wb),
            process_codes=process_codes,
            equipment_codes=equipment_codes,
            material_codes=material_codes,
            electricity_pending_years=pending,
        )
    finally:
        wb.close()


def gwp_lookup(data: SeedData) -> dict[str, float]:
    """轉成 calculator.derive_fuel_factor 要的 {氣體名稱: GWP} 形式。"""
    return {row.gas_name: row.gwp100 for row in data.gwp}
